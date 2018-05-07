import os
import openpyxl
from textblob import TextBlob

def sentanalysis():
    os.chdir('C:\\Users\\sabbas6\\Desktop\\Python')
    os.getcwd()

    wb = openpyxl.load_workbook('MRCS_KMI_Client_201705.xlsx', data_only = True)
    sheet = wb.active
    print(sheet)

    comments = [sheet['FU%s'%i].value for i in range(2,10)] ##sheet.max_row
    count = len(comments)
    print('Number of comments:',count,'\n')
    for x in range(len(comments)):
        textb = TextBlob(comments[x])
        tag = textb.tags
        word = textb.words
        print(comments[x])
        pol = textb.sentiment.polarity
        print('Polarity:',pol)
        if pol > 0:
            print('Positive')
        elif pol < 0:
            print('Negative')
        elif pol == 0:
            print('Neutral')
        subj = textb.sentiment.subjectivity
        print('Subjectivity:',subj)
        if subj > 0:
            print('Objective','\n')
        elif subj < 0:
            print('Subjective','\n')
        elif subj == 0:
            print('Neutral','\n')

sentanalysis()               
            





##for row in sheet.iter_rows('FU{}:FU{}'.format(sheet.min_row,sheet.max_row)):
##    #comments [] = sheet['FU2'].value
##    for cell in row:
##        comments.append(cell.value)
